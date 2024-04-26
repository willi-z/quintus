from ..datawriter import DataWriter

from quintus.helpers.id_generation import generate_id
from quintus.structures import Component, Composition, ValidationError, Measurement
from quintus.helpers.parser import parse_value
from quintus.composers import Composer

from .configuration import ExcelConfiguration, update_config, ExcelSheetConfiguration
from .legend import ExcelValueLegend

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import json
from pathlib import Path
import warnings
import io
import logging
logger = logging.getLogger(__name__)


class ExcelReader:
    def __init__(self, filename: str, config_file: str, writer: DataWriter, composers:set[Composer]= None):  # : DataWriter
        """Reads an Excel sheet with Data of ifferent components.
        Components can be grouped via the sheet name
        Values are color coded by a legend at the top left corner
        A pointers are used to find the table head and values


        Parameters
        ----------
        filename : str
            _description_
        config_file : str
            _description_
        writer : DataWriter
            _description_

        """
        self.filename = filename
        with Path(config_file).open() as fp:
            config = json.load(fp)

        try:
            self.config = ExcelConfiguration(**config)
        except ValidationError as ex:
            print(ex)

        self.writer = writer

        self.composers = composers

    def get_config(self) -> ExcelConfiguration:
        return self.config

    def read_all(self):
        with open(self.filename, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        wb = load_workbook(filename=in_mem_file, data_only=True, read_only=True)
        sheets_names = wb.sheetnames
        for sheet_name in sheets_names:
            if self.config.ignore is not None:
                if sheet_name in self.config.ignore:
                    continue

            if sheet_name not in self.config.sheets:
                continue

            master_config = self.config.model_dump()
            master_config.pop("ignore")
            master_config.pop("sheets")
            sheet_config = self.config.sheets.get(sheet_name).model_dump()
            self.read_sheet(wb, sheet_name, update_config(master_config, sheet_config))
        wb.close()

    def read_prefixes(
        self, sheet: Worksheet, config: ExcelSheetConfiguration
    ) -> list[str]:
        prefixes = []
        prefix_row = config.pointers.prefix
        for row in sheet.iter_rows(prefix_row, prefix_row):  # should only be one
            for cell in row:
                prefixes.append(cell.value)
        return prefixes

    def read_names(
        self, sheet: Worksheet, config: ExcelSheetConfiguration
    ) -> list[str]:
        names = []
        names_row = config.pointers.names
        for row in sheet.iter_rows(names_row, names_row):  # should only be one
            for cell in row:
                names.append(cell.value)
        return names

    def read_units(
        self, sheet: Worksheet, config: ExcelSheetConfiguration
    ) -> list[str]:
        units = []
        units_row = config.pointers.units
        for row in sheet.iter_rows(units_row, units_row):  # should only be one
            for cell in row:
                units.append(cell.value)
        return units

    def read_sheet(self, wb: Workbook, name: str, config: dict):
        sheet = wb[name]
        config = ExcelSheetConfiguration(**config)
        legend = ExcelValueLegend(sheet)
        prefix = self.read_prefixes(sheet, config)
        names = self.read_names(sheet, config)
        units = self.read_units(sheet, config)

        start_row = config.pointers.start
        for row in sheet.iter_rows(start_row):
            component = Component(identifier=generate_id())

            if component.tags is None:
                component.tags = set()

            if config.flag is not None:
                component.tags.add(config.flag.flag_as)

            for i in range(len(row)):
                cell = row[i]
                if cell.value is None:
                    continue

                cell_name = names[i]
                cell_unit = units[i]
                cell_prefix = prefix[i]

                if cell_name is None:
                    cell_name = names[i - 1]

                if cell_name.lower() in {"sources", "comment"}:
                    continue

                entry = component
                if cell_prefix is not None:
                    if not cell_prefix.startswith("{"):
                        if component.composition is None:
                            found_suitable_composer = False
                            if self.composers is not None:
                                for composer in self.composers:
                                    if composer.is_compatible_with(cell_prefix):
                                        found_suitable_composer = True
                                        component.composition = composer.generate(dict())
                                        logger.info(f"Detected compatible composer '{component.composition.type}' in row: '{cell.row}' (sheet: '{name}')")
                                        break
                            if not found_suitable_composer:
                                component.composition = Composition(components=dict())
                                logger.warning(f"Detected no compatible composer in row: '{cell.row}' (sheet: '{name}', column: '{cell.column}', prefix: '{cell_prefix}')")
                        if cell_prefix not in component.composition.components.keys():
                            entry = Component(identifier=generate_id())
                            component.composition.components[cell_prefix] = entry
                        else:
                            entry = component.composition.components[cell_prefix]
                        if cell_name == "material":
                            cell_name = "name"

                if cell.value is None:
                    continue
                if cell.value in {"-", "#DIV/0!"}:
                    continue

                if cell_name == "name":
                    entry.name = cell.value
                elif cell_name == "description":
                    entry.description = cell.value
                else:
                    value = {"value": cell.value}
                    if isinstance(cell.value, str):
                        if "+/-" in cell.value:
                            cell_val, tol = parse_value(cell.value)
                            value = {"value": cell_val, "tol": tol}
                    if cell_unit is not None:
                        value.update({"unit": cell_unit})
                    if cell_prefix is not None:
                        if cell_prefix.startswith("{"):  # is a json
                            measured_at = json.loads(cell_prefix)
                            value.update({"at": measured_at})

                    source = legend.get_value_type(cell)
                    if source is not None:
                        value.update({"source": source})

                    if entry.properties is None:
                        entry.properties = dict()
                    if isinstance(value["value"], str):
                        print(
                            f"Measurment {cell_name} from {entry.name} is not valid."
                            + f" (found in cell '{cell.coordinate}' "
                            + f"in sheet: '{sheet.title}')"
                        )
                    else:
                        try:
                            entry.properties[
                                cell_name.replace("  ", " ").replace(" ", "_")
                            ] = Measurement(**value)
                        except ValidationError:
                            warnings.warn(
                                f"Measurment {cell_name} from {entry.name}" 
                                + " is not valid."
                                + f" (found in cell '{cell.coordinate}' "
                                + f"in sheet: '{sheet.title}')"
                            )

            warnings.filterwarnings("error")
            try:
                component.warn_if_not_valid()
            except RuntimeWarning as warn:
                detail = ""
                if cell.data_type != "n":
                    detail = (
                        f" (found in cell '{cell.coordinate}' "
                        + f"in sheet: '{sheet.title}')"
                    )
                msg = warn.args[0] + detail
                warnings.resetwarnings()
                warnings.warn(msg)

            component.clear_empty()
            if component.is_valid():
                self.writer.write_entry(component)
            else:
                warnings.warn(
                    "After clearing empty entries the component is still not valid!"
                )
