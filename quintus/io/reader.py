from openpyxl import load_workbook
import json


def read_sheet(sheet, config: dict):
    for row in sheet.values:
        row_vals = []
        for value in row:
            row_vals.append(value)
        print(row_vals, sep=" | ")
        print()


def read_excel(excel_filename, config_file):
    wb = load_workbook(filename=excel_filename)
    config = json.load(config_file)
    sheets_names = wb.sheetnames
    for sheet_name in sheets_names:
        sheet = wb[sheet_name]
        # print(sheet.title)
        print(sheet_name)
        read_sheet(sheet, config)


if __name__ == "__main__":
    read_excel("tests/io/test_set1.xlsx")
